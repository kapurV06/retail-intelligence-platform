"""
Excel Report Generator — Professional boardroom-ready formatting.
Uses openpyxl with custom styles, conditional formatting, charts.
"""

import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from datetime import datetime
import os

# ── THEME ─────────────────────────────────────────────────────────────────────
DARK_NAVY    = "0D1B2A"
ACCENT_BLUE  = "1565C0"
ACCENT_TEAL  = "00897B"
ACCENT_RED   = "C62828"
ACCENT_AMBER = "F57F17"
LIGHT_GREY   = "F5F5F5"
MID_GREY     = "E0E0E0"
WHITE        = "FFFFFF"
TEXT_DARK    = "212121"


def _border(style="thin"):
    s = Side(style=style, color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)

def _header_fill(color=DARK_NAVY):
    return PatternFill("solid", fgColor=color)

def _header_font(bold=True, size=11, color=WHITE):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def _cell_font(bold=False, size=10, color=TEXT_DARK):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _fmt_currency(ws, cell_range):
    for row in ws[cell_range]:
        for cell in row:
            cell.number_format = '₹#,##0'

def _fmt_pct(ws, cell_range):
    for row in ws[cell_range]:
        for cell in row:
            cell.number_format = '0.00"%"'

def style_dataframe_sheet(ws, df, title, header_color=DARK_NAVY, freeze="A2"):
    """Write a DataFrame to a worksheet with full styling."""
    # Title row
    ws.append([title])
    ws.merge_cells(f"A1:{get_column_letter(len(df.columns))}1")
    title_cell = ws["A1"]
    title_cell.fill  = _header_fill(header_color)
    title_cell.font  = Font(name="Calibri", bold=True, size=14, color=WHITE)
    title_cell.alignment = _center()
    ws.row_dimensions[1].height = 28

    # Header row
    ws.append(list(df.columns))
    header_row = ws[2]
    for cell in header_row:
        cell.fill      = _header_fill("1E3A5F")
        cell.font      = _header_font(size=10)
        cell.alignment = _center()
        cell.border    = _border()
    ws.row_dimensions[2].height = 22

    # Data rows
    for i, row in enumerate(dataframe_to_rows(df, index=False, header=False)):
        ws.append(row)
        row_idx = i + 3
        fill_color = LIGHT_GREY if i % 2 == 0 else WHITE
        for j, cell in enumerate(ws[row_idx]):
            cell.fill      = PatternFill("solid", fgColor=fill_color)
            cell.font      = _cell_font()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = _border()
        ws.row_dimensions[row_idx].height = 18

    # Auto-width columns (skip merged cells)
    for col in ws.columns:
        try:
            col_letter = col[0].column_letter
            max_len = max((len(str(cell.value or "")) for cell in col if hasattr(cell, 'column_letter')), default=10)
            ws.column_dimensions[col_letter].width = min(max_len + 4, 35)
        except (AttributeError, TypeError):
            continue

    if freeze:
        ws.freeze_panes = freeze


def build_excel_report(eda_results, sql_results, output_path="../outputs/Retail_Intelligence_Report.xlsx"):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── SHEET 1: EXECUTIVE SUMMARY ────────────────────────────────────────────
    ws_exec = wb.create_sheet("📊 Executive Summary")
    ws_exec.sheet_view.showGridLines = False
    ws_exec.column_dimensions["A"].width = 3
    ws_exec.column_dimensions["B"].width = 30
    ws_exec.column_dimensions["C"].width = 22
    ws_exec.column_dimensions["D"].width = 22
    ws_exec.column_dimensions["E"].width = 22
    ws_exec.column_dimensions["F"].width = 3

    # Banner
    ws_exec.merge_cells("B1:E1")
    ws_exec["B1"] = "🛒 RETAIL INTELLIGENCE REPORT"
    ws_exec["B1"].fill = _header_fill(DARK_NAVY)
    ws_exec["B1"].font = Font(name="Calibri", bold=True, size=18, color=WHITE)
    ws_exec["B1"].alignment = _center()
    ws_exec.row_dimensions[1].height = 40

    ws_exec.merge_cells("B2:E2")
    ws_exec["B2"] = f"Inventory Stockout & Revenue Loss Analysis  |  Generated: {datetime.now().strftime('%d %b %Y')}"
    ws_exec["B2"].fill = _header_fill(ACCENT_BLUE)
    ws_exec["B2"].font = Font(name="Calibri", size=11, color=WHITE)
    ws_exec["B2"].alignment = _center()
    ws_exec.row_dimensions[2].height = 22

    # KPI Cards
    kpis = eda_results["kpis"]
    kpi_data = [
        ("💰 Total Revenue",         f"₹{kpis['total_revenue']:,.0f}",      ACCENT_BLUE,  "Revenue earned"),
        ("💀 Revenue Lost",           f"₹{kpis['total_lost_revenue']:,.0f}", ACCENT_RED,   "Lost to stockouts"),
        ("📉 Stockout Rate",          f"{kpis['stockout_rate']}%",           ACCENT_AMBER, "Of all demand events"),
        ("📦 Units Lost",             f"{kpis['total_units_lost']:,}",        ACCENT_TEAL,  "Units never sold"),
    ]
    kpi_more = [
        ("🎯 Gross Profit",           f"₹{kpis['total_gross_profit']:,.0f}", ACCENT_TEAL,  "After COGS"),
        ("🔻 Loss % of Potential",    f"{kpis['lost_revenue_pct']}%",        ACCENT_RED,   "Revenue at risk"),
        ("🛍️  Avg Daily Revenue",     f"₹{kpis['avg_daily_revenue']:,.0f}", ACCENT_BLUE,  "Per day average"),
        ("🏷️  Avg Discount Applied",  f"{kpis['avg_discount']}%",            ACCENT_AMBER, "When discounting"),
    ]

    def write_kpi_block(ws, row, kpi_list):
        cols = ["B","C","D","E"]
        for i, (label, value, color, sub) in enumerate(kpi_list):
            col = cols[i]
            ws.merge_cells(f"{col}{row}:{col}{row+1}")
            ws[f"{col}{row}"] = label
            ws[f"{col}{row}"].fill = _header_fill(color)
            ws[f"{col}{row}"].font = Font(name="Calibri", bold=True, size=10, color=WHITE)
            ws[f"{col}{row}"].alignment = _center()

            ws.merge_cells(f"{col}{row+2}:{col}{row+3}")
            ws[f"{col}{row+2}"] = value
            ws[f"{col}{row+2}"].fill = _header_fill("1E3A5F")
            ws[f"{col}{row+2}"].font = Font(name="Calibri", bold=True, size=14, color=WHITE)
            ws[f"{col}{row+2}"].alignment = _center()

            ws.merge_cells(f"{col}{row+4}:{col}{row+4}")
            ws[f"{col}{row+4}"] = sub
            ws[f"{col}{row+4}"].fill = PatternFill("solid", fgColor=LIGHT_GREY)
            ws[f"{col}{row+4}"].font = Font(name="Calibri", size=9, color="757575")
            ws[f"{col}{row+4}"].alignment = _center()

        for r in range(row, row+5):
            ws.row_dimensions[r].height = 18

    write_kpi_block(ws_exec, 4, kpi_data)
    write_kpi_block(ws_exec, 10, kpi_more)

    # Monthly table (mini)
    ws_exec["B17"] = "MONTHLY REVENUE SNAPSHOT"
    ws_exec.merge_cells("B17:E17")
    ws_exec["B17"].fill = _header_fill("1E3A5F")
    ws_exec["B17"].font = _header_font(size=11)
    ws_exec["B17"].alignment = _center()
    ws_exec.row_dimensions[17].height = 22

    monthly_mini = eda_results["monthly_trend"][["month","revenue","lost_revenue","stockout_rate"]].tail(6)
    monthly_mini.columns = ["Month","Revenue (₹)","Lost Revenue (₹)","Stockout %"]
    for j, col_name in enumerate(monthly_mini.columns):
        cell = ws_exec.cell(18, j+2, col_name)
        cell.fill = _header_fill(ACCENT_BLUE)
        cell.font = _header_font(size=10)
        cell.alignment = _center()
    for i, row in enumerate(monthly_mini.itertuples(index=False)):
        for j, val in enumerate(row):
            cell = ws_exec.cell(19+i, j+2, val)
            cell.fill = PatternFill("solid", fgColor=LIGHT_GREY if i%2==0 else WHITE)
            cell.font = _cell_font()
            cell.alignment = _center()
            cell.border = _border()

    # ── SHEET 2: CATEGORY PERFORMANCE ────────────────────────────────────────
    ws_cat = wb.create_sheet("📦 Category Analysis")
    ws_cat.sheet_view.showGridLines = False
    cat_df = eda_results["category_performance"][
        ["category","revenue","lost_revenue","gross_profit",
         "units_sold","stockout_rate","lost_rev_pct"]
    ].copy()
    cat_df.columns = ["Category","Revenue","Lost Revenue","Gross Profit",
                      "Units Sold","Stockout %","Loss %"]
    for col in ["Revenue","Lost Revenue","Gross Profit"]:
        cat_df[col] = cat_df[col].round(0).astype(int)
    style_dataframe_sheet(ws_cat, cat_df, "CATEGORY PERFORMANCE ANALYSIS", ACCENT_TEAL)

    # Conditional formatting: Loss % column (G)
    ws_cat.conditional_formatting.add(
        f"G3:G{len(cat_df)+2}",
        ColorScaleRule(start_type='min', start_color='63BE7B',
                       mid_type='percentile', mid_value=50, mid_color='FFEB84',
                       end_type='max', end_color='F8696B')
    )

    # ── SHEET 3: STORE LEADERBOARD ────────────────────────────────────────────
    ws_store = wb.create_sheet("🏪 Store Leaderboard")
    ws_store.sheet_view.showGridLines = False
    store_df = eda_results["store_performance"][
        ["revenue_rank","store_name","city","region","tier",
         "revenue","lost_revenue","stockout_rate","target_attainment"]
    ].head(25).copy()
    store_df.columns = ["Rank","Store Name","City","Region","Tier",
                        "Revenue","Lost Revenue","Stockout %","Target %"]
    for col in ["Revenue","Lost Revenue"]:
        store_df[col] = store_df[col].round(0).astype(int)
    style_dataframe_sheet(ws_store, store_df, "STORE PERFORMANCE LEADERBOARD", ACCENT_AMBER, freeze="A3")

    # Data bars on Revenue column
    ws_store.conditional_formatting.add(
        f"F3:F{len(store_df)+2}",
        DataBarRule(start_type='min', end_type='max',
                    color=ACCENT_BLUE, showValue=True)
    )

    # ── SHEET 4: CHRONIC STOCKOUTS (SQL) ──────────────────────────────────────
    ws_risk = wb.create_sheet("🚨 Stockout Risk")
    ws_risk.sheet_view.showGridLines = False
    risk_df = sql_results["chronic_stockouts"].copy()
    if not risk_df.empty:
        style_dataframe_sheet(ws_risk, risk_df, "HIGH-PRIORITY CHRONIC STOCKOUT SKUs", ACCENT_RED)
        ws_risk.conditional_formatting.add(
            f"G3:G{len(risk_df)+2}",
            ColorScaleRule(start_type='min', start_color='63BE7B',
                           end_type='max', end_color='F8696B')
        )

    # ── SHEET 5: SQL REVENUE RECOVERY ─────────────────────────────────────────
    ws_sql = wb.create_sheet("💡 Revenue Recovery")
    ws_sql.sheet_view.showGridLines = False
    style_dataframe_sheet(ws_sql, sql_results["revenue_recovery"],
                          "REVENUE RECOVERY OPPORTUNITY BY CATEGORY", ACCENT_TEAL)

    # ── SHEET 6: REGION × TIER MATRIX ─────────────────────────────────────────
    ws_matrix = wb.create_sheet("🗺️ Region Matrix")
    ws_matrix.sheet_view.showGridLines = False
    style_dataframe_sheet(ws_matrix, sql_results["region_tier_matrix"],
                          "REGION × STORE TIER REVENUE MATRIX", DARK_NAVY)

    # ── SHEET 7: RAW DATA (transactions sample) ────────────────────────────────
    ws_raw = wb.create_sheet("📋 Raw Data Sample")
    ws_raw.sheet_view.showGridLines = False

    wb.save(output_path)
    print(f"\n✅ Excel report saved: {output_path}")
    print(f"   Sheets: {[s.title for s in wb.worksheets]}")
